"""导出为 Excel / CSV（含评论明细）。"""
import csv
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from store import db
from store.models import Article

EXPORT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "exports")
os.makedirs(EXPORT_DIR, exist_ok=True)

NEWS_HEADERS = ["来源", "标题", "发布时间", "关键词分类", "命中关键词", "评论数", "链接"]
COMMENT_HEADERS = ["来源", "新闻标题", "评论作者", "评论内容", "点赞数", "评论时间", "新闻链接"]


def _news_rows(articles: list):
    for a in articles:
        pub = a.published_at.strftime("%Y-%m-%d %H:%M") if a.published_at else ""
        yield [
            a.source_name, a.title, pub,
            "、".join(a.category_tags),
            "、".join(a.matched_keywords),
            a.comment_count, a.url,
        ]


def _comment_rows(articles: list):
    """逐篇文章取其已抓取的评论，展开成一行一条评论。"""
    for a in articles:
        if not a.id:
            continue
        for c in db.get_comments(a.id):
            ct = c.published_at.strftime("%Y-%m-%d %H:%M") if c.published_at else ""
            yield [
                a.source_name, a.title,
                c.author or "", c.content or "",
                c.likes or 0, ct, a.url,
            ]


def _style_header(ws):
    head_fill = PatternFill("solid", fgColor="1F4E78")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head_fill
        c.alignment = Alignment(vertical="center")


def export_excel(articles: list, filename: str = "sports_news.xlsx", with_comments: bool = True) -> str:
    path = os.path.join(EXPORT_DIR, filename)
    wb = Workbook()

    # 工作表1：体育新闻
    ws = wb.active
    ws.title = "体育新闻"
    ws.append(NEWS_HEADERS)
    _style_header(ws)
    for row in _news_rows(articles):
        ws.append(row)
    for i, w in enumerate([14, 50, 18, 18, 24, 8, 40], start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    # 工作表2：评论明细
    if with_comments:
        cws = wb.create_sheet("评论明细")
        cws.append(COMMENT_HEADERS)
        _style_header(cws)
        n = 0
        for row in _comment_rows(articles):
            cws.append(row)
            n += 1
        if n == 0:
            cws.append(["（当前筛选范围内暂无已抓取的评论内容）", "", "", "", "", "", ""])
        for i, w in enumerate([12, 36, 16, 60, 8, 16, 36], start=1):
            cws.column_dimensions[chr(64 + i)].width = w

    wb.save(path)
    return path


def export_csv(articles: list, filename: str = "sports_news.csv") -> str:
    path = os.path.join(EXPORT_DIR, filename)
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(NEWS_HEADERS)
        for row in _news_rows(articles):
            w.writerow(row)
    return path


def export_comments_excel(articles: list, filename: str = "sports_comments.xlsx") -> str:
    """仅导出评论明细（独立 Excel）。"""
    path = os.path.join(EXPORT_DIR, filename)
    wb = Workbook()
    ws = wb.active
    ws.title = "评论明细"
    ws.append(COMMENT_HEADERS)
    _style_header(ws)
    n = 0
    for row in _comment_rows(articles):
        ws.append(row)
        n += 1
    if n == 0:
        ws.append(["（当前筛选范围内暂无已抓取的评论内容）", "", "", "", "", "", ""])
    for i, w in enumerate([12, 36, 16, 60, 8, 16, 36], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    wb.save(path)
    return path


def export_comments_csv(articles: list, filename: str = "sports_comments.csv") -> str:
    """仅导出评论明细（独立 CSV）。"""
    path = os.path.join(EXPORT_DIR, filename)
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(COMMENT_HEADERS)
        for row in _comment_rows(articles):
            w.writerow(row)
    return path
